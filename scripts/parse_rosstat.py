"""Парсер еженедельного XLSX Росстата для проекта «Что стало дороже?»."""
from pathlib import Path
import json
import sys
from openpyxl import load_workbook

MAPPING = {
    "gasoline95": {"name": "Бензин АИ-95", "shortName": "Бензин", "unit": "л", "sourceName": "Бензин автомобильный марки АИ-95, л"},
    "eggs": {"name": "Яйца куриные", "shortName": "Яйца", "unit": "10 шт.", "sourceName": "Яйца куриные, 10 шт."},
    "bread": {"name": "Хлеб пшеничный", "shortName": "Хлеб", "unit": "кг", "sourceName": "Хлеб и булочные изделия из пшеничной муки различных сортов, кг"},
    "milk": {"name": "Молоко пастеризованное", "shortName": "Молоко", "unit": "л", "sourceName": "Молоко питьевое цельное пастеризованное 2,5-3,2% жирности, л"},
    "potato": {"name": "Картофель", "shortName": "Картофель", "unit": "кг", "sourceName": "Картофель, кг"},
    "bus": {"name": "Проезд в городском автобусе", "shortName": "Автобус", "unit": "поездка", "sourceName": "Проезд в городском автобусе, поездка"},
}

MONTHS = {
    "января": "01", "февраля": "02", "марта": "03", "апреля": "04",
    "мая": "05", "июня": "06", "июля": "07", "августа": "08",
    "сентября": "09", "октября": "10", "ноября": "11", "декабря": "12",
}


def parse_header_date(text, year):
    text = str(text).replace("**", "").strip()
    if text.startswith("на "):
        text = text[3:]
    parts = text.split()
    if len(parts) < 2 or parts[1] not in MONTHS:
        raise ValueError(f"Не удалось распознать дату в заголовке: {text!r}")
    return f"{year}-{MONTHS[parts[1]]}-{int(parts[0]):02d}"


def choose_year_sheet(workbook):
    year_sheets = [name for name in workbook.sheetnames if str(name).isdigit() and len(str(name)) == 4]
    if not year_sheets:
        raise ValueError("В книге не найден лист с годом, например 2026")
    sheet_name = max(year_sheets, key=int)
    return workbook[sheet_name], int(sheet_name)


def main(src, dst):
    workbook = load_workbook(src, data_only=True)
    worksheet, year = choose_year_sheet(workbook)

    dates = [parse_header_date(cell.value, year) for cell in worksheet[4][1:] if cell.value]
    if len(dates) < 2:
        raise ValueError("В таблице найдено слишком мало недельных дат")

    by_name = {
        str(row[0].value).strip(): [cell.value for cell in row[1:1 + len(dates)]]
        for row in worksheet.iter_rows(min_row=5)
        if row[0].value
    }

    items = []
    for item_id, meta in MAPPING.items():
        source_name = meta["sourceName"]
        if source_name not in by_name:
            raise KeyError(f"Не найден показатель: {source_name}")
        values = by_name[source_name]
        if any(value is None for value in values):
            raise ValueError(f"В ряду {source_name!r} есть пустые значения")
        items.append({"id": item_id, **meta, "indices": [float(value) for value in values]})

    output = {
        "source": "Росстат",
        "sourceUrl": "https://rosstat.gov.ru/statistics/price",
        "scope": "Российская Федерация",
        "updated": dates[-1],
        "methodology": "Индексы потребительских цен в % к предыдущей дате регистрации. 7 дней = последняя недельная точка. 30 дней = накопленное изменение за последние 4 недельных интервала (28 дней).",
        "dates": dates,
        "items": items,
    }

    destination = Path(dst)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Готово: {destination} · актуально на {dates[-1]}")
    return output


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Использование: python scripts/parse_rosstat.py source/nedel_Ipc.xlsx data/prices.json")
        raise SystemExit(1)
    main(sys.argv[1], sys.argv[2])
